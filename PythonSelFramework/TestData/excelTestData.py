import openpyxl

book = openpyxl.load_workbook("F:\\Practice\\Pythonpractice\\PythonSelFramework\\TestData\\TestData.xlsx")
sheet = book.active
Dict = {}
# cell = sheet.cell(row=1,column=1)
# print(cell.value)

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "TC_1":
        for j in range(2, sheet.max_column+1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)
