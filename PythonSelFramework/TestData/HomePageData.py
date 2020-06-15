import openpyxl


class HomePageData:

    test_Homepage_data = \
        [{"firstname":"Abhishek", "lastname":"Singh", "pwd":"abc123", "gender":"Male", "DOB":"01-01-1993"},
                    {"firstname":"Swati", "lastname":"Tiwari", "pwd":"def123", "gender":"Female", "DOB":"01-02-1993"},
                    {"firstname":"Shruti", "lastname":"Deepak", "pwd":"ghi123", "gender":"Female", "DOB":"01-03-1993"}]

    @staticmethod
    def getTestData(testCaseName):
        Dict = {}
        book = openpyxl.load_workbook("F:\\Practice\\Pythonpractice\\PythonSelFramework\\TestData\\TestData.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testCaseName:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]
